
"""
Retail Date-Coding Automation Script
Processes raw scan export files into a structured, formatted Excel report.

Usage:
    python Datecoding_Supabarn.py -i Datecoding_Input.xlsx -o Datecoding_Output.xlsx
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
import argparse
import sys


def process_file(input_file: str, output_file: str):
    try:
        df = pd.read_excel(input_file, dtype={'Barcode': str})
    except FileNotFoundError:
        print(f"[✘] Error: Input file '{input_file}' not found.")
        sys.exit(1)

    # Rename Talker Count/Count to Date
    if 'Talker Count' in df.columns:
        df = df.rename(columns={'Talker Count': 'Date'})
    elif 'Count' in df.columns:
        df = df.rename(columns={'Count': 'Date'})
    else:
        raise KeyError("Column 'Talker Count' or 'Count' not found in input file")

    # Ensure Date is numeric & sort by Commodity Name → Date
    df['Date'] = pd.to_numeric(df['Date'], errors='coerce')
    df_sorted = df.sort_values(by=['Commodity Name', 'Date'])

    # Build final DataFrame with 2 blank rows after each commodity group
    rows = []
    for commodity, group in df_sorted.groupby('Commodity Name', sort=False):
        rows.append(group)
        blank = pd.DataFrame(
            [[""] * (len(df.columns) + 3)] * 2,
            columns=list(df.columns) + ['Today', 'Tomorrow', 'Gone']
        )
        rows.append(blank)
    final_df = pd.concat(rows, ignore_index=True)

    # Ensure Today / Tomorrow / Gone columns exist
    for col in ['Today', 'Tomorrow', 'Gone']:
        if col not in final_df.columns:
            final_df[col] = ""

    # Save to Excel
    final_df.to_excel(output_file, index=False)

    # Open with openpyxl for formatting
    wb = load_workbook(output_file)
    ws = wb.active

    # Auto-fit column widths
    for column_cells in ws.columns:
        max_len = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in column_cells
        )
        ws.column_dimensions[column_cells[0].column_letter].width = max_len + 2

    # Create styled Excel table
    last_row = ws.max_row
    last_col = ws.max_column
    table_range = f"A1:{ws.cell(row=last_row, column=last_col).coordinate}"

    tab = Table(displayName="DatecodingTable", ref=table_range)
    style = TableStyleInfo(
        name="TableStyleMedium22",
        showRowStripes=True,
        showColumnStripes=False
    )
    tab.tableStyleInfo = style
    ws.add_table(tab)

    wb.save(output_file)
    print(f"[✔] Report generated successfully: {output_file}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Retail Date-Coding Automation")
    parser.add_argument(
        "-i", "--input",
        default="Datecoding_Input.xlsx",
        help="Path to input Excel file"
    )
    parser.add_argument(
        "-o", "--output",
        default="Datecoding_Output.xlsx",
        help="Path to save output Excel file"
    )
    args = parser.parse_args()

    process_file(args.input, args.output)
